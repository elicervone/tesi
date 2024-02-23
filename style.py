from openpyxl.styles import Font, PatternFill, Border, Side


def style_sheet(sheet, num):
    ''' Funzione che crea lo stile del foglio Excel '''

    current_row = 14

    sheet.column_dimensions['A'].width = 75
    sheet.row_dimensions[1].height = 18
    sheet.column_dimensions['B'].width = 10
    

    sheet["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    sheet["A1"].font = Font(name='Arial', size=14, bold=True)
    sheet["A1"].border = Border(left=Side(border_style="thick", color="000000"),
                                right=Side(border_style="thick", color="000000"),
                                top=Side(border_style="thick", color="000000"),
                                bottom=Side(border_style="thick", color="000000"))

    for i in range(1, num+1):
        sheet.row_dimensions[current_row].height = 18
        sheet[f"A{current_row}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sheet[f"A{current_row}"].font = Font(name='Arial', size=12, bold=True)
        sheet[f"A{current_row}"].border = Border(left=Side(border_style="thick", color="000000"),
                                                    right=Side(border_style="thick", color="000000"),
                                                    top=Side(border_style="thick", color="000000"),
                                                    bottom=Side(border_style="thick", color="000000"))
        current_row += 7